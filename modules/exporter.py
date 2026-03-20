"""
exporter.py
Export attendance data to Excel (.xlsx) with formatting.
"""

import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, GradientFill)
from openpyxl.utils import get_column_letter
from database import db


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def export_daily(target_date: str = None, folder: str = None) -> str:
    """Export attendance for a single day. Returns saved file path."""
    if target_date is None:
        target_date = date.today().isoformat()
    if folder is None:
        folder = os.path.join(os.path.dirname(os.path.dirname(__file__)), "exports")
    os.makedirs(folder, exist_ok=True)

    records  = db.get_attendance_by_date(target_date)
    students = db.get_all_students()
    present_names = {r["name"] for r in records}

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # ── Title row ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Face Attendance Report — {target_date}"
    ws["A1"].font      = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="1A472A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Summary row ────────────────────────────────────────────────────────
    total   = len(students)
    present = len(records)
    absent  = total - present
    rate    = f"{round(present/total*100)}%" if total else "0%"

    ws.merge_cells("A2:F2")
    ws["A2"] = (f"Total: {total}   |   Present: {present}"
                f"   |   Absent: {absent}   |   Rate: {rate}")
    ws["A2"].font      = Font(name="Calibri", size=10, color="FFFFFF")
    ws["A2"].fill      = PatternFill("solid", fgColor="2D6A4F")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # ── Headers ────────────────────────────────────────────────────────────
    headers = ["#", "Name", "Roll No", "Department", "Time Marked", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font      = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="40916C")
        cell.alignment = Alignment(horizontal="center")
        cell.border    = _thin_border()
    ws.row_dimensions[3].height = 18

    # ── Data rows — all students, mark P/A ────────────────────────────────
    green_fill = PatternFill("solid", fgColor="D8F3DC")
    red_fill   = PatternFill("solid", fgColor="FFE8E8")

    for i, s in enumerate(students, 1):
        is_present = s["name"] in present_names
        time_str   = next((r["time"] for r in records if r["name"] == s["name"]), "—")
        status     = "Present" if is_present else "Absent"
        fill       = green_fill if is_present else red_fill
        row_data   = [i, s["name"], s["roll_no"] or "—",
                      s["department"] or "—", time_str, status]
        row_num    = i + 3

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill      = fill
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left")
            cell.border    = _thin_border()

    # ── Column widths ──────────────────────────────────────────────────────
    widths = [5, 28, 14, 18, 14, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    filepath = os.path.join(folder, f"attendance_{target_date}.xlsx")
    wb.save(filepath)
    return filepath


def export_range(from_date: str, to_date: str, folder: str = None) -> str:
    """Export attendance for a date range. Returns saved file path."""
    if folder is None:
        folder = os.path.join(os.path.dirname(os.path.dirname(__file__)), "exports")
    os.makedirs(folder, exist_ok=True)

    records = db.get_attendance_range(from_date, to_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Range Report"

    ws.merge_cells("A1:G1")
    ws["A1"] = f"Attendance Report: {from_date} to {to_date}"
    ws["A1"].font      = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="1A472A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["#", "Date", "Name", "Roll No", "Department", "Time", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font   = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill   = PatternFill("solid", fgColor="40916C")
        cell.alignment = Alignment(horizontal="center")
        cell.border = _thin_border()

    green_fill = PatternFill("solid", fgColor="D8F3DC")
    for i, r in enumerate(records, 1):
        row_data = [i, r["date"], r["name"], r["roll_no"] or "—",
                    r["department"] or "—", r["time"], r["status"]]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 2, column=col, value=val)
            cell.fill      = green_fill
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="center" if col != 3 else "left")
            cell.border    = _thin_border()

    widths = [5, 14, 28, 14, 18, 10, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    filepath = os.path.join(folder, f"report_{from_date}_to_{to_date}.xlsx")
    wb.save(filepath)
    return filepath
